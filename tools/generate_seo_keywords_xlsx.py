"""
Генератор Excel-файла с ключевыми запросами по разделам lenta-stalnaja.ru.

Собирает: 1 главную + 20 марок + 6 групп + 6 статических/утилитарных страниц.
На каждый раздел — 5 ключей с указанным интентом (коммерческий/информационный)
и приоритетом (high/medium/low). Столбец «Частота (WS)» оставлен пустым —
руководство заполнит после сбора семантики в Wordstat/KeyCollector.

Запуск: python tools/generate_seo_keywords_xlsx.py
Результат: export/seo-keywords.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── Данные ────────────────────────────────────────────────────────────────────

# Каждая запись: (Раздел, URL, [(ключ, intent, priority), ...])
SECTIONS = [
    (
        "Главная",
        "/",
        [
            ("стальная лента купить", "commercial", "high"),
            ("лента нержавеющая купить", "commercial", "high"),
            ("производитель стальной ленты", "commercial", "high"),
            ("лента стальная цена", "commercial", "high"),
            ("лента стальная оптом", "commercial", "medium"),
        ],
    ),

    # ── Марки стали ─────────────────────────────────────────────────────────
    ("Марка 12Х18Н10Т", "/12kh18n10t/", [
        ("лента 12Х18Н10Т", "commercial", "high"),
        ("лента 12Х18Н10Т купить", "commercial", "high"),
        ("лента 12Х18Н10Т цена", "commercial", "high"),
        ("12Х18Н10Т аналог AISI 321", "informational", "medium"),
        ("лента 12Х18Н10Т ГОСТ 4986-79", "informational", "medium"),
    ]),
    ("Марка 08Х18Н10", "/08kh18n10/", [
        ("лента 08Х18Н10", "commercial", "high"),
        ("лента 08Х18Н10 купить", "commercial", "high"),
        ("08Х18Н10 аналог AISI 304", "informational", "medium"),
        ("лента 08Х18Н10 характеристики", "informational", "medium"),
        ("лента 08Х18Н10 цена", "commercial", "high"),
    ]),
    ("Марка 10Х17Н13М3Т", "/10kh17n13m3t/", [
        ("лента 10Х17Н13М3Т", "commercial", "high"),
        ("лента 10Х17Н13М3Т купить", "commercial", "high"),
        ("10Х17Н13М3Т AISI 316Ti", "informational", "medium"),
        ("лента 10Х17Н13М3Т цена", "commercial", "high"),
        ("10Х17Н13М3Т характеристики", "informational", "medium"),
    ]),
    ("Марка 12Х18Н9", "/12kh18n9/", [
        ("лента 12Х18Н9", "commercial", "high"),
        ("лента 12Х18Н9 купить", "commercial", "high"),
        ("12Х18Н9 аналог AISI 302", "informational", "medium"),
        ("лента 12Х18Н9 характеристики", "informational", "medium"),
        ("лента 12Х18Н9 цена", "commercial", "medium"),
    ]),
    ("Марка 12Х18Н9СМР", "/12kh18n9smr/", [
        ("лента 12Х18Н9СМР", "commercial", "high"),
        ("12Х18Н9СМР ГОСТ 21996-76", "informational", "medium"),
        ("лента 12Х18Н9СМР купить", "commercial", "high"),
        ("12Х18Н9СМР характеристики", "informational", "low"),
        ("лента 12Х18Н9СМР цена", "commercial", "medium"),
    ]),
    ("Марка 17ХНГТ", "/17khngt/", [
        ("лента 17ХНГТ", "commercial", "high"),
        ("лента 17ХНГТ купить", "commercial", "high"),
        ("17ХНГТ ЭИ814", "informational", "medium"),
        ("лента 17ХНГТ ГОСТ 14117-85", "informational", "medium"),
        ("лента 17ХНГТ цена", "commercial", "medium"),
    ]),
    ("Марка 20Х13", "/20kh13/", [
        ("лента 20Х13", "commercial", "high"),
        ("лента 20Х13 купить", "commercial", "high"),
        ("20Х13 аналог AISI 420", "informational", "medium"),
        ("лента 20Х13 характеристики", "informational", "medium"),
        ("лента 20Х13 цена", "commercial", "high"),
    ]),
    ("Марка 27КХ", "/27kkh/", [
        ("лента 27КХ", "commercial", "high"),
        ("лента 27КХ купить", "commercial", "high"),
        ("27КХ прецизионный сплав", "informational", "medium"),
        ("27КХ сплав для магнитов", "informational", "medium"),
        ("лента 27КХ цена", "commercial", "medium"),
    ]),
    ("Марка 29НК", "/29nk/", [
        ("лента 29НК", "commercial", "high"),
        ("29НК Ковар", "informational", "medium"),
        ("лента 29НК купить", "commercial", "high"),
        ("29НК спай со стеклом", "informational", "medium"),
        ("лента 29НК цена", "commercial", "medium"),
    ]),
    ("Марка 36НХТЮ", "/36nkhtyu/", [
        ("лента 36НХТЮ", "commercial", "high"),
        ("лента 36НХТЮ купить", "commercial", "high"),
        ("36НХТЮ характеристики", "informational", "medium"),
        ("упругий сплав 36НХТЮ", "informational", "medium"),
        ("лента 36НХТЮ цена", "commercial", "medium"),
    ]),
    ("Марка 40КХНМ", "/40kkhnm/", [
        ("лента 40КХНМ", "commercial", "high"),
        ("лента 40КХНМ купить", "commercial", "high"),
        ("40КХНМ Havar аналог", "informational", "low"),
        ("лента 40КХНМ цена", "commercial", "medium"),
        ("40КХНМ характеристики", "informational", "medium"),
    ]),
    ("Марка 65Г", "/65g/", [
        ("лента 65Г", "commercial", "high"),
        ("пружинная лента 65Г", "commercial", "high"),
        ("лента 65Г купить", "commercial", "high"),
        ("лента 65Г цена", "commercial", "high"),
        ("лента 65Г ГОСТ 2283-79", "informational", "medium"),
    ]),
    ("Марка Х15Н60", "/kh15n60/", [
        ("лента Х15Н60", "commercial", "high"),
        ("лента Х15Н60 купить", "commercial", "high"),
        ("нихром Х15Н60 лента", "commercial", "medium"),
        ("Х15Н60 нагреватель", "informational", "medium"),
        ("лента Х15Н60 цена", "commercial", "medium"),
    ]),
    ("Марка Х15Ю5", "/kh15yu5/", [
        ("лента Х15Ю5", "commercial", "high"),
        ("фехраль Х15Ю5 лента", "commercial", "high"),
        ("Х15Ю5 купить", "commercial", "high"),
        ("лента Х15Ю5 характеристики", "informational", "medium"),
        ("лента Х15Ю5 цена", "commercial", "medium"),
    ]),
    ("Марка Х20Н80", "/kh20n80/", [
        ("лента Х20Н80", "commercial", "high"),
        ("нихром Х20Н80 лента", "commercial", "high"),
        ("лента Х20Н80 купить", "commercial", "high"),
        ("лента Х20Н80 цена", "commercial", "high"),
        ("Х20Н80 ГОСТ 12766", "informational", "medium"),
    ]),
    ("Марка Х20Н80-Н", "/kh20n80-n/", [
        ("лента Х20Н80-Н", "commercial", "high"),
        ("нихром нагартованный Х20Н80-Н", "informational", "medium"),
        ("лента Х20Н80-Н купить", "commercial", "high"),
        ("Х20Н80-Н характеристики", "informational", "medium"),
        ("лента Х20Н80-Н цена", "commercial", "medium"),
    ]),
    ("Марка Х23Ю5", "/kh23yu5/", [
        ("лента Х23Ю5", "commercial", "high"),
        ("фехраль Х23Ю5 лента", "commercial", "high"),
        ("Х23Ю5 Kanthal A", "informational", "medium"),
        ("лента Х23Ю5 купить", "commercial", "high"),
        ("лента Х23Ю5 цена", "commercial", "medium"),
    ]),
    ("Марка Х23Ю5Т", "/kh23yu5t/", [
        ("лента Х23Ю5Т", "commercial", "high"),
        ("фехраль Х23Ю5Т", "commercial", "high"),
        ("Х23Ю5Т купить", "commercial", "high"),
        ("лента Х23Ю5Т цена", "commercial", "medium"),
        ("Х23Ю5Т Kanthal AF", "informational", "medium"),
    ]),
    ("Марка ХН78Т", "/khn78t/", [
        ("лента ХН78Т", "commercial", "high"),
        ("лента ХН78Т купить", "commercial", "high"),
        ("ХН78Т Inconel 600 аналог", "informational", "medium"),
        ("жаропрочная лента ХН78Т", "commercial", "high"),
        ("лента ХН78Т цена", "commercial", "medium"),
    ]),
    ("Марка ЭИ814 (17ХНГТ)", "/ei814-17khngt/", [
        ("лента ЭИ814", "commercial", "high"),
        ("ЭИ814 17ХНГТ", "informational", "medium"),
        ("лента ЭИ814 купить", "commercial", "high"),
        ("ЭИ814 характеристики", "informational", "medium"),
        ("лента ЭИ814 цена", "commercial", "medium"),
    ]),

    # ── Группы по назначению ────────────────────────────────────────────────
    ("Группа: Коррозионно-стойкие стали", "/korrozionno-stojkie-stali/", [
        ("нержавеющая лента", "commercial", "high"),
        ("нержавеющая лента купить", "commercial", "high"),
        ("холоднокатаная нержавеющая лента", "commercial", "high"),
        ("нержавеющая лента ГОСТ 4986-79", "informational", "medium"),
        ("прайс нержавеющая лента", "commercial", "medium"),
    ]),
    ("Группа: Углеродистые стали", "/uglerodistye-stali/", [
        ("углеродистая лента", "commercial", "high"),
        ("пружинная лента", "commercial", "high"),
        ("лента 65Г купить", "commercial", "high"),
        ("пружинная сталь лента", "commercial", "medium"),
        ("лента углеродистая ГОСТ 2283-79", "informational", "medium"),
    ]),
    ("Группа: Высокое электросопротивление", "/vysokoe-elektrosoprotivlenie/", [
        ("нихромовая лента", "commercial", "high"),
        ("фехралевая лента", "commercial", "high"),
        ("лента для нагревателей", "commercial", "high"),
        ("лента для ТЭН", "commercial", "medium"),
        ("Х20Н80 лента купить", "commercial", "high"),
    ]),
    ("Группа: Прецизионные сплавы", "/precizionnye-splavy/", [
        ("прецизионные сплавы лента", "commercial", "high"),
        ("лента Ковар 29НК", "commercial", "medium"),
        ("прецизионная лента купить", "commercial", "high"),
        ("лента 36НХТЮ упругий сплав", "commercial", "medium"),
        ("прецизионная лента цена", "commercial", "medium"),
    ]),
    ("Группа: Жаростойкие и жаропрочные", "/zharostojkie-i-zharoprochnye/", [
        ("жаростойкая лента", "commercial", "high"),
        ("жаропрочная лента", "commercial", "high"),
        ("лента ХН78Т купить", "commercial", "high"),
        ("жаростойкая лента цена", "commercial", "medium"),
        ("Инконель лента аналог", "informational", "medium"),
    ]),
    ("Группа: Лента холоднокатаная", "/lenta-holodnokatanaya/", [
        ("холоднокатаная лента", "commercial", "high"),
        ("холоднокатаная лента ГОСТ", "informational", "medium"),
        ("лента холоднокатаная купить", "commercial", "high"),
        ("лента холоднокатаная стальная", "commercial", "high"),
        ("холоднокатаная лента цена", "commercial", "high"),
    ]),

    # ── Утилитарные / статические ───────────────────────────────────────────
    ("Калькулятор веса ленты", "/kalkulyator-vesa-lenty/", [
        ("калькулятор веса ленты", "informational", "high"),
        ("расчёт веса нержавеющей ленты", "informational", "high"),
        ("сколько весит стальная лента", "informational", "medium"),
        ("калькулятор массы ленты онлайн", "informational", "medium"),
        ("вес рулона стальной ленты", "informational", "medium"),
    ]),
    ("Справочник ГОСТов", "/gost/", [
        ("ГОСТ 4986-79 лента нержавеющая", "informational", "high"),
        ("ГОСТ 14117-85 прецизионные сплавы", "informational", "medium"),
        ("ГОСТ 2283-79 лента 65Г", "informational", "medium"),
        ("ГОСТ на нихромовую ленту", "informational", "medium"),
        ("стандарты стальной ленты", "informational", "medium"),
    ]),
    ("О компании", "/about/", [
        ("производитель стальной ленты", "commercial", "high"),
        ("поставщик металлопроката", "commercial", "high"),
        ("оптовый поставщик ленты", "commercial", "medium"),
        ("лента стальная нижний новгород", "commercial", "high"),
        ("купить ленту оптом производитель", "commercial", "medium"),
    ]),
    ("Доставка", "/delivery/", [
        ("доставка стальной ленты по России", "commercial", "high"),
        ("стальная лента доставка транспортной", "commercial", "medium"),
        ("купить ленту с доставкой", "commercial", "high"),
        ("доставка металлопроката ТК", "commercial", "medium"),
        ("доставка нержавеющей ленты Москва СПб", "commercial", "high"),
    ]),
    ("Контакты", "/contacts/", [
        ("купить ленту нижний новгород", "commercial", "high"),
        ("купить ленту москва", "commercial", "high"),
        ("купить ленту СПб", "commercial", "high"),
        ("стальная лента телефон менеджер", "commercial", "medium"),
        ("стальная лента заказать", "commercial", "high"),
    ]),
    ("Оплата", "/payment/", [
        ("лента стальная для юрлиц", "commercial", "high"),
        ("нержавеющая лента с НДС", "commercial", "high"),
        ("лента стальная безналичный расчёт", "commercial", "medium"),
        ("лента стальная счёт-фактура", "commercial", "medium"),
        ("оплата стальной ленты по счёту", "commercial", "medium"),
    ]),
]

# ── Построение файла ──────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Семантика lenta-stalnaja"

# Заголовок и стили
HEADER = ["№", "Раздел", "URL", "Ключевой запрос", "Intent", "Приоритет", "Частота (WS)", "Комментарий"]

ws.append(HEADER)

# Стили заголовка
header_font = Font(bold=True, color="FFFFFF", size=11, name="Inter")
header_fill = PatternFill("solid", fgColor="1D4ED8")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="94A3B8")
header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx in range(1, len(HEADER) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = header_border

# Стили строк
priority_fills = {
    "high":   PatternFill("solid", fgColor="DBEAFE"),  # синий
    "medium": PatternFill("solid", fgColor="FEF3C7"),  # жёлтый
    "low":    PatternFill("solid", fgColor="F1F5F9"),  # серый
}
section_font = Font(bold=True, size=10, name="Inter")
regular_font = Font(size=10, name="Inter")
mono_font = Font(size=10, name="Consolas")

row = 2
n = 1
last_section = None
for section, url, keys in SECTIONS:
    for (kw, intent, priority) in keys:
        ws.cell(row=row, column=1, value=n).font = regular_font
        # Раздел показываем только у первой строки группы — для читаемости
        section_cell = ws.cell(row=row, column=2, value=section if section != last_section else "")
        section_cell.font = section_font if section != last_section else regular_font
        url_cell = ws.cell(row=row, column=3, value=url if section != last_section else "")
        url_cell.font = mono_font
        url_cell.hyperlink = f"https://lenta-stalnaja.ru{url}" if section != last_section else None
        if section != last_section:
            url_cell.font = Font(size=10, name="Consolas", color="2563EB", underline="single")

        kw_cell = ws.cell(row=row, column=4, value=kw)
        kw_cell.font = regular_font

        intent_cell = ws.cell(row=row, column=5, value="Коммерческий" if intent == "commercial" else "Информационный")
        intent_cell.font = regular_font
        intent_cell.alignment = Alignment(horizontal="center")

        priority_cell = ws.cell(row=row, column=6, value={"high": "Высокий", "medium": "Средний", "low": "Низкий"}[priority])
        priority_cell.font = regular_font
        priority_cell.alignment = Alignment(horizontal="center")
        priority_cell.fill = priority_fills[priority]

        # Столбцы 7 и 8 — пустые для ручного заполнения
        ws.cell(row=row, column=7, value="").font = regular_font
        ws.cell(row=row, column=8, value="").font = regular_font

        # Общая тонкая рамка
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = Border(
                left=Side(border_style="thin", color="E2E5E9"),
                right=Side(border_style="thin", color="E2E5E9"),
                top=Side(border_style="thin", color="E2E5E9"),
                bottom=Side(border_style="thin", color="E2E5E9"),
            )

        last_section = section
        row += 1
        n += 1

# Ширины столбцов
widths = {
    "A": 5,    # №
    "B": 32,   # Раздел
    "C": 32,   # URL
    "D": 42,   # Ключ
    "E": 16,   # Intent
    "F": 12,   # Приоритет
    "G": 14,   # Частота
    "H": 30,   # Комментарий
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Заморозка первой строки + автофильтр
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{row - 1}"

# Высота заголовка
ws.row_dimensions[1].height = 34

# Сохраняем
out_dir = Path(__file__).resolve().parent.parent / "export"
out_dir.mkdir(parents=True, exist_ok=True)
out_path = out_dir / "seo-keywords.xlsx"
wb.save(out_path)
print(f"OK -> {out_path}")
print(f"Строк: {row - 2}, разделов: {len(SECTIONS)}")
